import io
import logging
import uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response, Query, status
from sqlalchemy.orm import Session
import openpyxl

from app.api.deps import get_current_user
from app.core.database import get_db
from app.models.user import User
from app.models.reference_data import InternationalMarketBrand, RegisteredNotInUse, WhoInnRegistry
from app.repositories.screening import ScreeningRepository
from app.repositories.settings import SettingsRepository
from app.schemas.reference_data import (
    DataSourceListResponse,
    DataSourceStatus,
    DataSourceToggleRequest,
    ReferenceDataStatus,
    UploadResponse,
    WhoInnRecordCreate,
    WhoInnRecordUpdate,
    WhoInnRecordResponse,
    InternationalMarketRecordCreate,
    InternationalMarketRecordUpdate,
    InternationalMarketRecordResponse,
    RegisteredNotInUseRecordCreate,
    RegisteredNotInUseRecordUpdate,
    RegisteredNotInUseRecordResponse,
)
from app.services.market_check import google_search_configured
from app.services.tabular_import import (
    normalize_name,
    parse_international_market_xlsx,
    parse_registered_not_in_use_xlsx,
)
from app.services.who_inn_import import normalize_inn_name, parse_who_inn_pdf

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reference-data", tags=["Reference Data"])

_XLSX_CONTENT_TYPES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
)


def _require_superuser(current_user: User) -> None:
    if not current_user.is_superuser and current_user.role not in ("super_admin", "admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")


def _require_xlsx(file: UploadFile) -> None:
    if file.content_type not in _XLSX_CONTENT_TYPES and not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File must be an .xlsx workbook")


@router.get("/status", response_model=ReferenceDataStatus)
def get_reference_data_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    repo = ScreeningRepository(db)
    return ReferenceDataStatus(
        who_inn_row_count=repo.who_inn_row_count(),
        iqvia_row_count=repo.iqvia_row_count(),
        registered_not_in_use_row_count=repo.registered_not_in_use_row_count(),
        international_market_row_count=repo.international_market_row_count(),
    )


# ---------------------------------------------------------------------------
# Toggleable data sources — Connect / Disconnect
# ---------------------------------------------------------------------------

def _build_data_source_statuses(repo: ScreeningRepository, toggles: dict) -> list[DataSourceStatus]:
    who_count = repo.who_inn_row_count()
    iqvia_count = repo.iqvia_row_count()
    who_enabled = toggles.get("who_inn_enabled", True)
    iqvia_enabled = toggles.get("iqvia_enabled", True)
    epharmacy_enabled = toggles.get("epharmacy_enabled", True)
    google_enabled = toggles.get("google_search_enabled", True)
    google_configured = google_search_configured()

    return [
        DataSourceStatus(
            id="who_inn",
            name="WHO INN Registry",
            category="Regulatory Databases",
            description="International Nonproprietary Names. A name identical to a registered INN is an automatic knockout.",
            enabled=who_enabled,
            connected=who_enabled,
            detail=(
                f"{who_count} entries loaded from upload." if who_count > 0
                else "No entries loaded, falling back to the live ChEMBL API per screen."
            ) if who_enabled else "Disabled. WHO INN is not checked during screening.",
        ),
        DataSourceStatus(
            id="iqvia",
            name="IQVIA Extract",
            category="Market Intelligence",
            description="Licensed market-data extract. Requires a confirmed IQVIA licence before any row counts as coverage.",
            enabled=iqvia_enabled,
            connected=iqvia_enabled and iqvia_count > 0,
            detail=(
                f"{iqvia_count} entries loaded." if iqvia_count > 0
                else "No data loaded, licence not yet confirmed."
            ) if iqvia_enabled else "Disabled. IQVIA is not checked during screening.",
        ),
        DataSourceStatus(
            id="epharmacy",
            name="E-Pharmacy Platforms",
            category="Market Intelligence",
            description="Live scrape of 1mg, PharmEasy, Apollo Pharmacy & Netmeds for active listings under the candidate name.",
            enabled=epharmacy_enabled,
            connected=epharmacy_enabled,
            detail=(
                "Live scrape active: 1mg, PharmEasy, Apollo Pharmacy, Netmeds." if epharmacy_enabled
                else "Disabled. E-pharmacy platforms are not checked during screening."
            ),
        ),
        DataSourceStatus(
            id="google_search",
            name="Google Search",
            category="Market Intelligence",
            description="General web presence check via Google Custom Search: market/regulatory mentions outside e-pharmacy listings.",
            enabled=google_enabled,
            connected=google_enabled and google_configured,
            detail=(
                ("Configured." if google_configured else "Not configured. Set GOOGLE_API_KEY/GOOGLE_CSE_ID.")
                if google_enabled else "Disabled. Google Search is not checked during screening."
            ),
        ),
    ]


@router.get("/data-sources", response_model=DataSourceListResponse)
def list_data_sources(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    repo = ScreeningRepository(db)
    toggles = SettingsRepository(db).get_data_source_toggles()
    return DataSourceListResponse(sources=_build_data_source_statuses(repo, toggles))


@router.put("/data-sources/{source_id}", response_model=DataSourceStatus)
def toggle_data_source(
    source_id: str,
    payload: DataSourceToggleRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    key_map = {
        "who_inn": "who_inn_enabled",
        "iqvia": "iqvia_enabled",
        "epharmacy": "epharmacy_enabled",
        "google_search": "google_search_enabled",
    }
    if source_id not in key_map:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Unknown data source '{source_id}'")

    settings_repo = SettingsRepository(db)
    toggles = settings_repo.set_data_source_toggle(key_map[source_id], payload.enabled)
    logger.info("[DATA SOURCE TOGGLE] %s set to enabled=%s by %s", source_id, payload.enabled, current_user.email)

    repo = ScreeningRepository(db)
    statuses = _build_data_source_statuses(repo, toggles)
    return next(s for s in statuses if s.id == source_id)


# ---------------------------------------------------------------------------
# WHO INN — PDF upload
# ---------------------------------------------------------------------------

@router.post("/who-inn/upload", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_who_inn_list(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)

    if file.content_type not in ("application/pdf", "application/octet-stream") and not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File must be a PDF")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")

    try:
        parsed = parse_who_inn_pdf(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception:
        logger.exception("Unexpected failure parsing WHO INN PDF upload %r", file.filename)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Could not read this PDF. Is it the expected WHO INN list template?")

    rows = [
        {"inn_name": r["inn_name"], "normalized_name": normalize_inn_name(r["inn_name"]),
         "who_publication_reference": r["who_publication_reference"]}
        for r in parsed
    ]

    repo = ScreeningRepository(db)
    imported = repo.replace_all_who_inn(rows)
    logger.info("[WHO INN IMPORT] Replaced who_inn_registry with %d rows from %r (by %s)",
                imported, file.filename, current_user.email)

    return UploadResponse(
        rows_imported=imported,
        message=f"Imported {imported} WHO INN entries from '{file.filename}'.",
    )


# ---------------------------------------------------------------------------
# MASTER DATA — WHO INN Registry (CRUD + Template + Export)
# ---------------------------------------------------------------------------

@router.get("/who-inn/records", response_model=list[WhoInnRecordResponse])
def get_who_inn_records(
    q: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(WhoInnRegistry)
    if q and q.strip():
        search_pattern = f"%{q.strip()}%"
        query = query.filter(
            (WhoInnRegistry.inn_name.ilike(search_pattern)) |
            (WhoInnRegistry.who_publication_reference.ilike(search_pattern))
        )
    return query.order_by(WhoInnRegistry.inn_name.asc()).limit(500).all()


@router.post("/who-inn/records", response_model=WhoInnRecordResponse, status_code=status.HTTP_201_CREATED)
def create_who_inn_record(
    payload: WhoInnRecordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    if not payload.inn_name or not payload.inn_name.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="INN name is required.")

    record = WhoInnRegistry(
        inn_name=payload.inn_name.strip(),
        normalized_name=normalize_inn_name(payload.inn_name),
        who_publication_reference=payload.who_publication_reference.strip() if payload.who_publication_reference else None,
        chembl_id=payload.chembl_id.strip() if payload.chembl_id else None,
        molecule_type=payload.molecule_type.strip() if payload.molecule_type else None,
        created_at=datetime.now(timezone.utc),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.put("/who-inn/records/{record_id}", response_model=WhoInnRecordResponse)
def update_who_inn_record(
    record_id: uuid.UUID,
    payload: WhoInnRecordUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    record = db.query(WhoInnRegistry).filter(WhoInnRegistry.id == record_id).first()
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Record not found.")

    if payload.inn_name is not None:
        if not payload.inn_name.strip():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="INN name cannot be empty.")
        record.inn_name = payload.inn_name.strip()
        record.normalized_name = normalize_inn_name(payload.inn_name)
    if payload.who_publication_reference is not None:
        record.who_publication_reference = payload.who_publication_reference.strip() if payload.who_publication_reference else None
    if payload.chembl_id is not None:
        record.chembl_id = payload.chembl_id.strip() if payload.chembl_id else None
    if payload.molecule_type is not None:
        record.molecule_type = payload.molecule_type.strip() if payload.molecule_type else None

    db.commit()
    db.refresh(record)
    return record


@router.delete("/who-inn/records/{record_id}", status_code=status.HTTP_200_OK)
def delete_who_inn_record(
    record_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    record = db.query(WhoInnRegistry).filter(WhoInnRegistry.id == record_id).first()
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Record not found.")
    db.delete(record)
    db.commit()
    return {"message": "Record deleted successfully."}


@router.get("/who-inn/template")
def download_who_inn_template(
    current_user: User = Depends(get_current_user),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WHO INN Template"
    headers = ["Sr No.", "INN Name", "W.H.O Publication Reference"]
    ws.append(headers)

    ws.append([1, "ABACAVIR", "List 77 (1997)"])
    ws.append([2, "METFORMIN", "List 4 (1956)"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=WHO_INN_Template.xlsx"},
    )


@router.get("/who-inn/export")
def export_who_inn_master_data(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    records = db.query(WhoInnRegistry).order_by(WhoInnRegistry.inn_name.asc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WHO INN Registry"
    headers = ["Sr No.", "INN Name", "W.H.O Publication Reference", "Created Date"]
    ws.append(headers)

    for idx, r in enumerate(records, start=1):
        ws.append([
            idx,
            r.inn_name,
            r.who_publication_reference or "",
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=WHO_INN_Registry_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


# ---------------------------------------------------------------------------
# MASTER DATA — International Market Brands (CRUD + Template + Export + Bulk)
# ---------------------------------------------------------------------------

@router.get("/international-market/records", response_model=list[InternationalMarketRecordResponse])
def get_international_market_records(
    q: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(InternationalMarketBrand)
    if q and q.strip():
        search_pattern = f"%{q.strip()}%"
        query = query.filter(
            (InternationalMarketBrand.brand_name.ilike(search_pattern)) |
            (InternationalMarketBrand.active_ingredient.ilike(search_pattern)) |
            (InternationalMarketBrand.country.ilike(search_pattern))
        )
    return query.order_by(InternationalMarketBrand.created_at.desc()).all()


@router.post("/international-market/records", response_model=InternationalMarketRecordResponse, status_code=status.HTTP_201_CREATED)
def create_international_market_record(
    payload: InternationalMarketRecordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    if not payload.brand_name or not payload.brand_name.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Brand name (Mark) is required.")

    record = InternationalMarketBrand(
        brand_name=payload.brand_name.strip(),
        normalized_name=normalize_name(payload.brand_name),
        active_ingredient=payload.active_ingredient.strip() if payload.active_ingredient else None,
        country=payload.country.strip() if payload.country else None,
        as_of_date=payload.as_of_date,
        created_at=datetime.now(timezone.utc),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.put("/international-market/records/{record_id}", response_model=InternationalMarketRecordResponse)
def update_international_market_record(
    record_id: uuid.UUID,
    payload: InternationalMarketRecordUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    record = db.query(InternationalMarketBrand).filter(InternationalMarketBrand.id == record_id).first()
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Record not found.")

    if payload.brand_name is not None:
        if not payload.brand_name.strip():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Brand name cannot be empty.")
        record.brand_name = payload.brand_name.strip()
        record.normalized_name = normalize_name(payload.brand_name)
    if payload.active_ingredient is not None:
        record.active_ingredient = payload.active_ingredient.strip() if payload.active_ingredient else None
    if payload.country is not None:
        record.country = payload.country.strip() if payload.country else None
    if payload.as_of_date is not None:
        record.as_of_date = payload.as_of_date

    db.commit()
    db.refresh(record)
    return record


@router.delete("/international-market/records/{record_id}", status_code=status.HTTP_200_OK)
def delete_international_market_record(
    record_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    record = db.query(InternationalMarketBrand).filter(InternationalMarketBrand.id == record_id).first()
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Record not found.")
    db.delete(record)
    db.commit()
    return {"message": "Record deleted successfully."}


@router.get("/international-market/template")
def download_international_market_template(
    current_user: User = Depends(get_current_user),
):
    """Generates the exact template matching info/data_sourrces_uploads/international.xlsx:
    Headers: Sr No., Mark, Molecule"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Sr No.", "Mark", "Molecule"])

    # Sample rows matching the official template
    ws.append([1, "TICAPLET", "TICAGRELOR"])
    ws.append([2, "TICAFAST", "TICAGRELOR"])
    ws.append([3, "ANGRELOR", "TICAGRELOR"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=International_Markets_Template.xlsx"},
    )


@router.get("/international-market/export")
def export_international_market_master_data(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    records = db.query(InternationalMarketBrand).order_by(InternationalMarketBrand.brand_name.asc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Sr No.", "Mark", "Molecule"])

    for idx, r in enumerate(records, start=1):
        ws.append([
            idx,
            r.brand_name,
            r.active_ingredient or "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=International_Market_Brands_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.post("/international-market/upload", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_international_market(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    _require_xlsx(file)

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")

    try:
        parsed = parse_international_market_xlsx(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception:
        logger.exception("Unexpected failure parsing International Market upload %r", file.filename)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Could not read this workbook. Is it the expected template?")

    repo = ScreeningRepository(db)
    imported = repo.replace_all_international_market(parsed)
    logger.info("[INTERNATIONAL-MARKET IMPORT] Replaced table with %d rows from %r (by %s)",
                imported, file.filename, current_user.email)

    return UploadResponse(
        rows_imported=imported,
        message=f"Imported {imported} International Market Brand entries from '{file.filename}'.",
    )


# ---------------------------------------------------------------------------
# MASTER DATA — Registered but Not in Use (CRUD + Template + Export + Bulk)
# ---------------------------------------------------------------------------

@router.get("/registered-not-in-use/records", response_model=list[RegisteredNotInUseRecordResponse])
def get_registered_not_in_use_records(
    q: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(RegisteredNotInUse)
    if q and q.strip():
        search_pattern = f"%{q.strip()}%"
        query = query.filter(
            (RegisteredNotInUse.brand_name.ilike(search_pattern)) |
            (RegisteredNotInUse.application_number.ilike(search_pattern)) |
            (RegisteredNotInUse.status.ilike(search_pattern)) |
            (RegisteredNotInUse.remarks.ilike(search_pattern))
        )
    return query.order_by(RegisteredNotInUse.created_at.desc()).all()


@router.post("/registered-not-in-use/records", response_model=RegisteredNotInUseRecordResponse, status_code=status.HTTP_201_CREATED)
def create_registered_not_in_use_record(
    payload: RegisteredNotInUseRecordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    if not payload.brand_name or not payload.brand_name.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Brand Name (TradeMark Name) is required.")

    record = RegisteredNotInUse(
        brand_name=payload.brand_name.strip(),
        normalized_name=normalize_name(payload.brand_name),
        trademark_class=payload.trademark_class or 5,
        application_number=payload.application_number.strip() if payload.application_number else None,
        application_date=payload.application_date,
        status=payload.status.strip() if payload.status else "Registered",
        valid_till=payload.valid_till,
        remarks=payload.remarks.strip() if payload.remarks else None,
        as_of_date=payload.as_of_date,
        created_at=datetime.now(timezone.utc),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.put("/registered-not-in-use/records/{record_id}", response_model=RegisteredNotInUseRecordResponse)
def update_registered_not_in_use_record(
    record_id: uuid.UUID,
    payload: RegisteredNotInUseRecordUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    record = db.query(RegisteredNotInUse).filter(RegisteredNotInUse.id == record_id).first()
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Record not found.")

    if payload.brand_name is not None:
        if not payload.brand_name.strip():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Brand name cannot be empty.")
        record.brand_name = payload.brand_name.strip()
        record.normalized_name = normalize_name(payload.brand_name)
    if payload.trademark_class is not None:
        record.trademark_class = payload.trademark_class
    if payload.application_number is not None:
        record.application_number = payload.application_number.strip() if payload.application_number else None
    if payload.application_date is not None:
        record.application_date = payload.application_date
    if payload.status is not None:
        record.status = payload.status.strip() if payload.status else None
    if payload.valid_till is not None:
        record.valid_till = payload.valid_till
    if payload.remarks is not None:
        record.remarks = payload.remarks.strip() if payload.remarks else None
    if payload.as_of_date is not None:
        record.as_of_date = payload.as_of_date

    db.commit()
    db.refresh(record)
    return record


@router.delete("/registered-not-in-use/records/{record_id}", status_code=status.HTTP_200_OK)
def delete_registered_not_in_use_record(
    record_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    record = db.query(RegisteredNotInUse).filter(RegisteredNotInUse.id == record_id).first()
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Record not found.")
    db.delete(record)
    db.commit()
    return {"message": "Record deleted successfully."}


@router.get("/registered-not-in-use/template")
def download_registered_not_in_use_template(
    current_user: User = Depends(get_current_user),
):
    """Generates the exact template matching info/data_sourrces_uploads/register_not_used.xlsx:
    Headers: Sl No, TradeMark Name, Class, Appl No, Appl Date, TMR STATUS, Valid till, Description, E-commerce URL"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["Sl No", "TradeMark Name", "Class", "Appl No", "Appl Date", "TMR STATUS", "Valid till", "Description", "E-commerce URL"]
    ws.append(headers)

    # Sample rows matching the official template
    ws.append([1, "A2CLEAR", 5, "5693827", "2022-11-22", "Registered", "2032-11-22", "N-1 Recommended to maintain this mark", ""])
    ws.append([2, "ACTRIL", 5, "564924", "1992-01-03", "Registered", "2026-01-03", "Dormant registered mark", ""])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Registered_Not_In_Use_Template.xlsx"},
    )


@router.get("/registered-not-in-use/export")
def export_registered_not_in_use_master_data(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    records = db.query(RegisteredNotInUse).order_by(RegisteredNotInUse.brand_name.asc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["Sl No", "TradeMark Name", "Class", "Appl No", "Appl Date", "TMR STATUS", "Valid till", "Description", "E-commerce URL"]
    ws.append(headers)

    for idx, r in enumerate(records, start=1):
        ws.append([
            idx,
            r.brand_name,
            r.trademark_class or 5,
            r.application_number or "",
            r.application_date.strftime("%Y-%m-%d") if r.application_date else "",
            r.status or "Registered",
            r.valid_till.strftime("%Y-%m-%d") if r.valid_till else "",
            r.remarks or "",
            "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Registered_Not_In_Use_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.post("/registered-not-in-use/upload", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_registered_not_in_use(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_superuser(current_user)
    _require_xlsx(file)

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")

    try:
        parsed = parse_registered_not_in_use_xlsx(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception:
        logger.exception("Unexpected failure parsing Registered-Not-in-Use upload %r", file.filename)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Could not read this workbook. Is it the expected template?")

    repo = ScreeningRepository(db)
    imported = repo.replace_all_registered_not_in_use(parsed)
    logger.info("[REGISTERED-NOT-IN-USE IMPORT] Replaced table with %d rows from %r (by %s)",
                imported, file.filename, current_user.email)

    return UploadResponse(
        rows_imported=imported,
        message=f"Imported {imported} Registered-but-Not-in-Use entries from '{file.filename}'.",
    )
