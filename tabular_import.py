"""Deterministic parsers for the two Excel-based Tier-1 reference imports:

  - "Registered but Not in Use" — the Trade Marks Registry's own export of
    registered-but-unused marks (e.g. "Un-used TradeMarks List 2026.xlsx").
    Matched against the real workbook used to build this: it ships with two
    sheets that use *different* column sets for the same data (one has a
    "Discontinued brand" free-text column and an unlabeled Y/N column, the
    other has "Description"/"E-commerce URL" instead) — headers are matched
    by name, not position, and both sheets are read.
  - "International Market Brands" — an overseas brand-name search export
    (e.g. "OVERSEAS BRAND-NAMES_SEARCHED.xlsx"): a flat 3-column sheet
    (Sr No., Mark, Molecule) with no country breakdown.

Both are template-specific (matched to the real files this importer
was built against), just tolerant of header order/casing/aliases within
that template.
"""
import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip()).lower()


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _build_column_index(header_row: Tuple[Any, ...], aliases: Dict[str, List[str]]) -> Dict[str, int]:
    col_index: Dict[str, int] = {}
    for idx, raw_header in enumerate(header_row):
        key = _normalize_header(raw_header)
        if not key:
            continue
        for field, alias_list in aliases.items():
            if field in col_index:
                continue
            if key in alias_list:
                col_index[field] = idx
    return col_index


def _cell(row: Tuple[Any, ...], col_index: Dict[str, int], field: str) -> Any:
    idx = col_index.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ---------------------------------------------------------------------------
# Registered but Not in Use
# ---------------------------------------------------------------------------

_REG_NOT_IN_USE_ALIASES: Dict[str, List[str]] = {
    "brand_name": ["trademarkname", "trademark", "unusedtrademarkname"],
    "trademark_class": ["class", "tmclass", "trademarkclass"],
    "application_number": ["applno", "applicationno", "applicationnumber"],
    "application_date": ["appldate", "applicationdate"],
    "status": ["tmrstatus", "tmstatus"],
    "valid_till": ["validtill", "validuntil", "expiry", "expirydate"],
    "remarks": ["discontinuedbrand", "description", "remarks", "notes"],
}


def parse_registered_not_in_use_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    deduped: Dict[Tuple[str, Optional[str]], Dict[str, Any]] = {}

    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            continue
        col_index = _build_column_index(header_row, _REG_NOT_IN_USE_ALIASES)
        
        # Strict validation: MUST contain trademark name and at least one TM specific field
        if "brand_name" not in col_index or (
            "trademark_class" not in col_index and "application_number" not in col_index and "status" not in col_index
        ):
            continue

        for row in rows_iter:
            brand_name_raw = _cell(row, col_index, "brand_name")
            if not brand_name_raw or not str(brand_name_raw).strip():
                continue
            brand_name = str(brand_name_raw).strip()

            class_raw = _cell(row, col_index, "trademark_class")
            trademark_class = None
            if class_raw is not None:
                try:
                    trademark_class = int(class_raw)
                except (TypeError, ValueError):
                    pass

            application_number = _cell(row, col_index, "application_number")
            application_number = str(application_number).strip() if application_number is not None else None

            app_date_raw = _cell(row, col_index, "application_date")
            application_date = app_date_raw if isinstance(app_date_raw, datetime) else None

            valid_till_raw = _cell(row, col_index, "valid_till")
            valid_till = valid_till_raw if isinstance(valid_till_raw, datetime) else None

            status_raw = _cell(row, col_index, "status")
            status = str(status_raw).strip() if status_raw is not None else None

            # Keep unparsed remarks
            remarks_parts = []
            remarks_raw = _cell(row, col_index, "remarks")
            if remarks_raw is not None and str(remarks_raw).strip():
                remarks_parts.append(str(remarks_raw).strip())
            if valid_till_raw is not None and valid_till is None:
                remarks_parts.append(f"Valid till (unparsed): {valid_till_raw}")

            key = (normalize_name(brand_name), application_number)
            deduped[key] = {
                "brand_name": brand_name,
                "normalized_name": normalize_name(brand_name),
                "trademark_class": trademark_class,
                "application_number": application_number,
                "application_date": application_date,
                "status": status,
                "valid_till": valid_till,
                "remarks": " | ".join(remarks_parts) or None,
            }

    if not deduped:
        raise ValueError(
            "This workbook does not match the 'Registered but Not in Use' template. "
            "(Expected columns: 'TradeMark Name', 'Class', 'Appl No', 'TMR STATUS')."
        )
    return list(deduped.values())


# ---------------------------------------------------------------------------
# International Market Brands
# ---------------------------------------------------------------------------

_INTL_MARKET_ALIASES: Dict[str, List[str]] = {
    "brand_name": ["mark", "internationalbrandname", "overseasbrandname"],
    "active_ingredient": ["molecule", "activeingredient", "genericname", "generic"],
    "country": ["country", "market"],
}


def parse_international_market_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    deduped: Dict[Tuple[str, Optional[str]], Dict[str, Any]] = {}

    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            continue
        col_index = _build_column_index(header_row, _INTL_MARKET_ALIASES)
        
        # Strict validation: MUST contain Mark AND Molecule
        if "brand_name" not in col_index or "active_ingredient" not in col_index:
            continue

        for row in rows_iter:
            brand_name_raw = _cell(row, col_index, "brand_name")
            if not brand_name_raw or not str(brand_name_raw).strip():
                continue
            brand_name = str(brand_name_raw).strip()

            ingredient_raw = _cell(row, col_index, "active_ingredient")
            active_ingredient = str(ingredient_raw).strip() if ingredient_raw is not None else None

            country_raw = _cell(row, col_index, "country")
            country = str(country_raw).strip() if country_raw is not None else None

            key = (normalize_name(brand_name), active_ingredient)
            deduped[key] = {
                "brand_name": brand_name,
                "normalized_name": normalize_name(brand_name),
                "active_ingredient": active_ingredient,
                "country": country,
            }

    if not deduped:
        raise ValueError(
            "This workbook does not match the 'International Market Brands' template. "
            "(Expected columns: 'Mark', 'Molecule')."
        )
    return list(deduped.values())
